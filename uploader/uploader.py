"""Asmama 크롤링 데이터를 Qoo10 업로드 형식으로 변환하는 메인 시스템."""

import os
import json
import logging
from pathlib import Path
from typing import Dict, Any, List, Optional
import pandas as pd
from datetime import datetime

from data_loader import TemplateLoader
from image_processor import ImageProcessor
from product_filter import ProductFilter
from field_transformer import FieldTransformer


class AsamaUploader:
    """
    Asmama 크롤링 데이터를 Qoo10 업로드 형식으로 변환하는 메인 클래스.
    
    전체 워크플로:
    1. 템플릿 파일 로딩
    2. 이미지 품질 검사 및 대표 이미지 선정
    3. 상품 필터링 (금지브랜드, 경고키워드, 기등록상품)
    4. 필드 변환 (번역, 가격변환, 카테고리매핑)
    5. Excel 파일 출력
    """
    
    def __init__(self, templates_dir: str, output_dir: str = "output", image_filter_mode: str = "advanced"):
        """
        AsamaUploader 초기화.
        
        Args:
            templates_dir: 템플릿 파일들이 있는 디렉토리
            output_dir: 출력 파일 저장 디렉토리
            image_filter_mode: 이미지 필터링 모드 ("ai", "advanced", "both")
        """
        self.templates_dir = Path(templates_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        # 로깅 설정
        self.logger = logging.getLogger(__name__)
        
        # 구성 요소 초기화
        self.template_loader = TemplateLoader(templates_dir)
        self.image_processor = ImageProcessor(filter_mode=image_filter_mode)
        self.product_filter = None  # template_loader 로딩 후 초기화
        self.field_transformer = None  # template_loader 로딩 후 초기화
        
        # 통계
        self.stats = {
            "total_input_products": 0,
            "image_processed_products": 0,
            "filtered_products": 0,
            "transformed_products": 0,
            "final_output_products": 0
        }
    
    def load_templates(self) -> bool:
        """
        템플릿 파일들을 로딩한다.
        
        Returns:
            로딩 성공 여부
        """
        try:
            success = self.template_loader.load_all_templates()
            if success:
                # 템플릿 로딩 후 필터링 및 변환 시스템 초기화
                self.product_filter = ProductFilter(self.template_loader)
                self.field_transformer = FieldTransformer(self.template_loader)
                self.logger.info("템플릿 로딩 및 시스템 초기화 완료")
            return success
        except Exception as e:
            self.logger.error(f"템플릿 로딩 실패: {str(e)}")
            return False
    
    def process_crawled_data(self, input_file: str) -> bool:
        """
        크롤링된 데이터를 처리하여 Qoo10 업로드 형식으로 변환한다.
        
        Args:
            input_file: 크롤링된 데이터 파일 경로 (Excel)
            
        Returns:
            처리 성공 여부
        """
        try:
            # 1. 입력 데이터 로딩
            products = self._load_crawled_data(input_file)
            if not products:
                self.logger.error("입력 데이터 로딩 실패")
                return False
            
            self.stats["total_input_products"] = len(products)
            self.logger.info(f"test 입력 데이터 로딩 완료: {len(products)}개 상품")
            
            # 2. 이미지 품질 검사 및 대표 이미지 선정
            image_processed_products = self._process_images(products)
            self.stats["image_processed_products"] = len(image_processed_products)
            
            # 3. 상품 필터링
            filtered_products, filter_stats = self.product_filter.filter_products(image_processed_products)
            self.stats["filtered_products"] = len(filtered_products)
            
            # 4. 필드 변환
            transformed_products = self.field_transformer.transform_products(filtered_products)
            self.stats["transformed_products"] = len(transformed_products)
            
            # 5. 최종 검증 및 Excel 출력
            output_success = self._save_to_excel(transformed_products)
            if output_success:
                self.stats["final_output_products"] = len(transformed_products)
            
            # 6. 결과 리포트 생성
            self._generate_report(filter_stats)
            
            return output_success
            
        except Exception as e:
            self.logger.error(f"데이터 처리 실패: {str(e)}")
            return False
    
    def _load_crawled_data(self, input_file: str) -> List[Dict[str, Any]]:
        """
        크롤링된 데이터를 로딩한다.
        
        Args:
            input_file: 입력 파일 경로
            
        Returns:
            상품 데이터 목록
        """
        try:
            input_path = Path(input_file)
            
            if input_path.suffix.lower() == '.xlsx':
                # Excel 파일 로딩
                df = pd.read_excel(input_path)
            elif input_path.suffix.lower() == '.json':
                # JSON 파일 로딩
                with open(input_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if isinstance(data, list):
                        df = pd.DataFrame(data)
                    else:
                        df = pd.DataFrame([data])
            else:
                self.logger.error(f"지원하지 않는 파일 형식: {input_path.suffix}")
                return []
            
            # DataFrame을 딕셔너리 목록으로 변환
            products = df.to_dict('records')
            
            # NaN 값을 빈 문자열로 변환
            for product in products:
                for key, value in product.items():
                    if pd.isna(value):
                        product[key] = ""
            
            self.logger.info(f"크롤링 데이터 로딩 완료: {len(products)}개 상품")
            return products
            
        except Exception as e:
            self.logger.error(f"크롤링 데이터 로딩 실패: {str(e)}")
            return []
    
    def _process_images(self, products: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        모든 상품의 이미지를 처리한다.
        
        Args:
            products: 상품 목록
            
        Returns:
            이미지 처리된 상품 목록
        """
        self.logger.info(f"이미지 처리 시작: {len(products)}개 상품")
        
        processed_products = []
        for i, product in enumerate(products, 1):
            try:
                processed_product = self.image_processor.process_product_images(product)
                processed_products.append(processed_product)
                
                if i % 10 == 0:
                    self.logger.info(f"이미지 처리 진행중: {i}/{len(products)}개 완료")
                    
            except Exception as e:
                self.logger.error(f"이미지 처리 실패: {product.get('branduid', 'unknown')} - {str(e)}")
                processed_products.append(product)  # 실패해도 원본 데이터 유지
        
        self.logger.info(f"이미지 처리 완료: {len(processed_products)}개 상품")
        return processed_products
    
    def _save_to_excel(self, products: List[Dict[str, Any]]) -> bool:
        """
        sample.xlsx 템플릿을 기반으로 변환된 데이터를 덮어써서 저장한다.
        
        Args:
            products: 변환된 상품 목록
            
        Returns:
            저장 성공 여부
        """
        try:
            if not products:
                self.logger.warning("저장할 상품 데이터가 없음")
                return False
            
            # sample.xlsx 템플릿 파일 경로
            sample_file = self.templates_dir / "upload" / "sample.xlsx"
            if not sample_file.exists():
                self.logger.error(f"샘플 템플릿 파일이 없음: {sample_file}")
                return False
            
            self.logger.info(f"샘플 템플릿 로딩: {sample_file}")
            
            # 샘플 템플릿 로딩 (Row 0을 컬럼명으로 사용)
            template_df = pd.read_excel(sample_file, header=0)  # Row 0을 헤더로 사용
            
            if template_df.empty:
                self.logger.error("샘플 템플릿이 비어있음")
                return False
            
            # 기존 데이터 모두 삭제 (헤더만 유지)
            template_df = template_df.iloc[0:0]  # 빈 DataFrame이지만 컬럼은 유지
            
            self.logger.info(f"템플릿 컬럼 수: {len(template_df.columns)}개")
            
            # 변환된 상품 데이터를 DataFrame으로 변환
            products_df = pd.DataFrame(products)
            
            # 템플릿의 모든 컬럼에 맞춰 데이터 정렬 및 누락된 컬럼 채움
            for col in template_df.columns:
                if col not in products_df.columns:
                    products_df[col] = ""  # 누락된 컬럼은 빈 값으로 채움
            
            # 템플릿 컬럼 순서로 정렬
            products_df = products_df[template_df.columns]
            
            # 템플릿에 새 데이터 추가
            final_df = pd.concat([template_df, products_df], ignore_index=True)
            
            # 출력 파일명 생성
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = self.output_dir / f"qoo10_upload_{timestamp}.xlsx"
            
            # 원본 sample.xlsx의 구조를 유지하면서 저장
            # openpyxl로 원본 파일을 열고 데이터 시트만 교체
            from openpyxl import load_workbook
            
            # 원본 샘플 파일 복사
            import shutil
            shutil.copy2(sample_file, output_file)
            
            # 복사된 파일에서 데이터 시트 업데이트
            workbook = load_workbook(output_file)
            
            # 첫 번째 워크시트 선택 (보통 데이터가 있는 시트)
            worksheet = workbook.active
            
            # 기존 데이터 영역 삭제 (Row 5부터, 원본 구조: Row 0=컬럼명, Row 1-3=설명, Row 4=빈행, Row 5부터=데이터)
            max_row = worksheet.max_row
            if max_row > 4:  # Row 5부터 삭제
                for row in worksheet.iter_rows(min_row=5, max_row=max_row):
                    for cell in row:
                        cell.value = None
            
            # 새로운 데이터 쓰기 (Row 5부터 시작)
            for row_idx, (_, row_data) in enumerate(final_df.iterrows(), start=5):
                for col_idx, value in enumerate(row_data, start=1):
                    worksheet.cell(row=row_idx, column=col_idx, value=value)
            
            # 파일 저장
            workbook.save(output_file)
            workbook.close()
            
            self.logger.info(f"샘플 템플릿 기반 Excel 파일 저장 완료: {output_file} ({len(products)}개 상품)")
            return True
            
        except Exception as e:
            self.logger.error(f"Excel 파일 저장 실패: {str(e)}")
            return False
    
    def _generate_report(self, filter_stats: Dict[str, Any]) -> None:
        """
        처리 결과 리포트를 생성한다.
        
        Args:
            filter_stats: 필터링 통계
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_file = self.output_dir / f"processing_report_{timestamp}.txt"
            
            # 리포트 내용 생성
            report_lines = []
            report_lines.append("🚀 Asmama → Qoo10 업로드 데이터 변환 리포트")
            report_lines.append("=" * 60)
            report_lines.append("")
            
            # 전체 통계
            report_lines.append("📊 전체 처리 통계:")
            report_lines.append(f"  입력 상품 수: {self.stats['total_input_products']:,}개")
            report_lines.append(f"  이미지 처리 완료: {self.stats['image_processed_products']:,}개")
            report_lines.append(f"  필터링 통과: {self.stats['filtered_products']:,}개")
            report_lines.append(f"  필드 변환 완료: {self.stats['transformed_products']:,}개")
            report_lines.append(f"  최종 출력: {self.stats['final_output_products']:,}개")
            
            success_rate = (self.stats['final_output_products'] / self.stats['total_input_products'] * 100) if self.stats['total_input_products'] > 0 else 0
            report_lines.append(f"  전체 성공률: {success_rate:.1f}%")
            report_lines.append("")
            
            # 필터링 상세 통계
            if filter_stats:
                filter_summary = self.product_filter.get_filter_summary(filter_stats)
                report_lines.append(filter_summary)
                report_lines.append("")
            
            # 변환 상세 통계
            transform_summary = self.field_transformer.get_transformation_summary(
                self.stats['filtered_products'], 
                self.stats['transformed_products']
            )
            report_lines.append(transform_summary)
            
            # 리포트 파일 저장
            with open(report_file, 'w', encoding='utf-8') as f:
                f.write("\n".join(report_lines))
            
            self.logger.info(f"처리 리포트 생성 완료: {report_file}")
            
            # 콘솔에도 요약 출력
            print("\n" + "=" * 60)
            print("🚀 Asmama → Qoo10 업로드 데이터 변환 완료!")
            print(f"📊 결과: {self.stats['total_input_products']:,}개 → {self.stats['final_output_products']:,}개 (성공률: {success_rate:.1f}%)")
            print("=" * 60)
            
        except Exception as e:
            self.logger.error(f"리포트 생성 실패: {str(e)}")


def main():
    """
    메인 실행 함수.
    """
    import argparse
    
    parser = argparse.ArgumentParser(description="Asmama 크롤링 데이터를 Qoo10 업로드 형식으로 변환")
    parser.add_argument("--input", required=True, help="크롤링 데이터 파일 경로 (Excel/JSON)")
    parser.add_argument("--templates", default="templates", help="템플릿 파일 디렉토리 (기본값: templates)")
    parser.add_argument("--output", default="output", help="출력 디렉토리 (기본값: output)")
    parser.add_argument("--log-level", default="INFO", choices=["DEBUG", "INFO", "WARNING", "ERROR"])
    parser.add_argument("--image-filter", default="advanced", choices=["ai", "advanced", "both"], 
                       help="이미지 필터링 모드 (기본값: advanced) - ai: OpenAI Vision API, advanced: 로직 필터링, both: 둘 다")
    
    args = parser.parse_args()
    
    # 로그 디렉토리 생성
    LOGS_DIR = Path("logs")
    LOGS_DIR.mkdir(exist_ok=True)
    
    # 로그 파일 핸들러
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_handler = logging.FileHandler(
        LOGS_DIR / f"uploader_{timestamp}.log",
        encoding='utf-8'
    )
    file_handler.setLevel(logging.DEBUG)  # 파일에는 DEBUG 레벨까지 저장
    
    # 콘솔 핸들러
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)  # 콘솔은 INFO만
    
    # 포맷터
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)
    
    # 로깅 설정
    logging.basicConfig(
        level=getattr(logging, args.log_level),
        handlers=[file_handler, console_handler],
        force=True
    )
    
    # 업로더 실행
    uploader = AsamaUploader(args.templates, args.output, args.image_filter)
    
    try:
        # 템플릿 로딩
        if not uploader.load_templates():
            print("❌ 템플릿 로딩 실패")
            return False
        
        # 데이터 처리
        success = uploader.process_crawled_data(args.input)
        
        if success:
            print("✅ 데이터 변환 성공!")
            return True
        else:
            print("❌ 데이터 변환 실패")
            return False
            
    except Exception as e:
        print(f"❌ 실행 오류: {str(e)}")
        return False


if __name__ == "__main__":
    main()